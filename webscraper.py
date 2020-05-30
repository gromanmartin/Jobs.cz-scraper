import requests
import numpy as np
from bs4 import BeautifulSoup
import openpyxl
from itertools import compress


class Scraper:

    def __init__(self, url):
        self.url = url
        self.page = requests.get(self.url)
        self.soup = BeautifulSoup(self.page.content, 'html.parser')
        self.keywords = ['data', 'analyst', 'science', 'analýza', 'business', 'consultant', 'machine learning'
                         'analytic']    # at least one of these words will be included in the results
        self.blacklist = ['senior', 'manager', 'manažer']    # no job titles including these words will be in the results

    def scan(self):
        page_results = self.soup.find_all('a', class_='search-list__main-info__title__link')
        page_links = [a['href'] for a in page_results]
        page_job_titles = [a.text.lower() for a in page_results]
        relevant_indexes = self.match_keywords(page_job_titles)
        final_index = self.find_index_of_yday()
        mask = [x < final_index for x in relevant_indexes]
        relevant_indexes = list(compress(relevant_indexes, mask))
        job_titles_results = [page_job_titles[i] for i in relevant_indexes]
        job_links_results = [page_links[i] for i in relevant_indexes]
        job_titles_results, job_links_results = self.check_blacklisted(job_titles_results, job_links_results)
        return job_links_results, job_titles_results

    def output_to_excel(self, titles, links):
        # Outputs the results into excel file with job titles and ad links
        df = zip(titles, links)
        filename = 'output.xlsx'
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        row = ws.max_row + 1
        for i, (title, link) in enumerate(df):
            ws.cell(row=row+i, column=1, value=title)
            ws.cell(row=row+i, column=2, value=link)
        wb.save(filename)

    def match_keywords(self, job_titles):
        # Returns a list of indexes of job titles, which contain any keyword
        old_list = np.array(len(job_titles) * [False])
        for keyword in self.keywords:
            l = np.array([keyword in job_title for job_title in job_titles])
            new_list = l | old_list
            old_list = list(new_list)
        indexes = [i for i, x in enumerate(old_list) if x]
        return indexes

    def find_index_of_yday(self):
        # Return the index of job advertisement added yesterday, or puts index = 30 if all ads are from today
        page_labels_added = self.soup.find_all('span', class_='label-added')
        labels = [x.text for x in page_labels_added]
        word = ['Přidáno včera', 'Aktualizováno včera']
        try:
            ind = labels.index(word[0])
            if ind == 30:
                ind = labels.index(word[1])
        except ValueError:
            ind = 30
        return ind

    def check_blacklisted(self, job_list, link_list):
        # returns lists without blacklisted words in job titles
        for blackword in self.blacklist:
            for (job, link) in zip(job_list, link_list):
                if blackword in job:
                    link_list.remove(link)
                    job_list.remove(job)
        return job_list, link_list


class PageGenerator:

    def __init__(self):
        self.pages = []

    def set_pages(self):
        # Returns a list of pages from selected date
        page_number = 1
        location = 'praha'
        while True:
            new_page = 'https://www.jobs.cz/prace/{}/?page={}&locality%5Bradius%5D=0'.format(location, page_number)
            url = requests.get(new_page)
            soup = BeautifulSoup(url.content, 'html.parser')
            self.pages.append(new_page)
            page_number += page_number
            if self.from_post(soup):
                break
        return self.pages

    def from_post(self, soup):
        # Returns true/false if current page contains post from date set in from_when
        results = soup.find_all('span', class_='label-added')
        labels = [x.text for x in results]
        from_when = ['Přidáno včera', 'Aktualizováno včera']
        output = any(word in labels for word in from_when)
        print(output, labels)
        return output


if __name__ == '__main__':
    pagegen = PageGenerator()
    urls = pagegen.set_pages()
    for page in urls:
        scraper = Scraper(page)
        links, titles = scraper.scan()
        scraper.output_to_excel(titles, links)
        print(titles)
